#!/usr/bin/env python3
"""
PowerNode Python Excel Analysis MCP Server
Advanced Excel analysis using openpyxl for features not available in ExcelJS

TOOLS:
1. extract_comments - Extract all Excel comments from a workbook
2. extract_questions - Find all cells containing questions (cells with '?')
3. detect_hidden_content - List all hidden rows and columns
4. extract_formulas - Extract all formulas from a workbook
5. read_with_hidden - Read data including hidden rows/columns
6. comprehensive_analysis - Full workbook analysis (all features)
"""

import json
import sys
import base64
import tempfile
import os
from pathlib import Path
from typing import Any, Dict, List
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Azure Storage imports (optional - for direct file access)
try:
    from azure.storage.blob import BlobServiceClient
    AZURE_AVAILABLE = True
except ImportError:
    AZURE_AVAILABLE = False

class ExcelAnalyzer:
    """Excel analysis using openpyxl"""

    @staticmethod
    def load_workbook_from_buffer(file_buffer: bytes):
        """Load workbook from buffer"""
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(file_buffer)
            tmp_path = tmp_file.name

        try:
            wb = load_workbook(tmp_path, data_only=False)
            return wb, tmp_path
        except Exception as e:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise e

    @staticmethod
    def extract_comments(filename: str, file_buffer: bytes = None) -> Dict[str, Any]:
        """Extract all comments from an Excel workbook"""
        wb, tmp_path = ExcelAnalyzer.load_workbook_from_buffer(file_buffer)

        try:
            results = {
                "filename": filename,
                "total_comments": 0,
                "worksheets": []
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_comments = []

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.comment:
                            sheet_comments.append({
                                "cell": cell.coordinate,
                                "value": str(cell.value) if cell.value else "",
                                "comment": cell.comment.text,
                                "author": getattr(cell.comment, 'author', 'Unknown')
                            })

                if sheet_comments:
                    results["worksheets"].append({
                        "sheet": sheet_name,
                        "comment_count": len(sheet_comments),
                        "comments": sheet_comments
                    })
                    results["total_comments"] += len(sheet_comments)

            wb.close()
            return results
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @staticmethod
    def extract_questions(filename: str, file_buffer: bytes = None) -> Dict[str, Any]:
        """Find all cells containing questions (cells with '?')"""
        wb, tmp_path = ExcelAnalyzer.load_workbook_from_buffer(file_buffer)

        try:
            results = {
                "filename": filename,
                "total_questions": 0,
                "worksheets": []
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_questions = []

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and '?' in str(cell.value):
                            # Try to find answer in adjacent cells (next cell to the right)
                            answer_cell = ws.cell(row=cell.row, column=cell.column + 1)

                            sheet_questions.append({
                                "cell": cell.coordinate,
                                "question": str(cell.value),
                                "answer_cell": answer_cell.coordinate if answer_cell else None,
                                "answer": str(answer_cell.value) if answer_cell and answer_cell.value else ""
                            })

                if sheet_questions:
                    results["worksheets"].append({
                        "sheet": sheet_name,
                        "question_count": len(sheet_questions),
                        "questions": sheet_questions
                    })
                    results["total_questions"] += len(sheet_questions)

            wb.close()
            return results
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @staticmethod
    def detect_hidden_content(filename: str, file_buffer: bytes = None) -> Dict[str, Any]:
        """List all hidden rows and columns"""
        wb, tmp_path = ExcelAnalyzer.load_workbook_from_buffer(file_buffer)

        try:
            results = {
                "filename": filename,
                "total_hidden_rows": 0,
                "total_hidden_columns": 0,
                "worksheets": []
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                hidden_rows = []
                hidden_columns = []

                # Check hidden rows
                for row_num in range(1, ws.max_row + 1):
                    row_dim = ws.row_dimensions[row_num]
                    if row_dim.hidden:
                        hidden_rows.append(row_num)

                # Check hidden columns
                for col_num in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_num)
                    col_dim = ws.column_dimensions[col_letter]
                    if col_dim.hidden:
                        hidden_columns.append(col_letter)

                sheet_data = {
                    "sheet": sheet_name,
                    "hidden": ws.sheet_state == 'hidden',
                    "row_count": ws.max_row,
                    "column_count": ws.max_column,
                    "hidden_row_count": len(hidden_rows),
                    "hidden_column_count": len(hidden_columns),
                    "hidden_rows": hidden_rows[:50],  # Limit to first 50
                    "hidden_columns": hidden_columns[:50]
                }

                results["worksheets"].append(sheet_data)
                results["total_hidden_rows"] += len(hidden_rows)
                results["total_hidden_columns"] += len(hidden_columns)

            wb.close()
            return results
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    @staticmethod
    def comprehensive_analysis(filename: str, file_buffer: bytes = None) -> Dict[str, Any]:
        """Full workbook analysis combining all features"""
        comments = ExcelAnalyzer.extract_comments(filename, file_buffer)
        questions = ExcelAnalyzer.extract_questions(filename, file_buffer)
        hidden = ExcelAnalyzer.detect_hidden_content(filename, file_buffer)

        return {
            "filename": filename,
            "summary": {
                "total_comments": comments["total_comments"],
                "total_questions": questions["total_questions"],
                "total_hidden_rows": hidden["total_hidden_rows"],
                "total_hidden_columns": hidden["total_hidden_columns"],
                "worksheet_count": len(hidden["worksheets"])
            },
            "comments": comments,
            "questions": questions,
            "hidden_content": hidden
        }

# MCP Server Implementation
TOOLS = [
    {
        "name": "extract_comments",
        "description": "Extract all Excel comments from a workbook. Returns comments with cell location, content, and author.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Workbook filename (.xlsx file)"
                }
            },
            "required": ["filename"]
        }
    },
    {
        "name": "extract_questions",
        "description": "Find all cells containing questions (cells with '?'). Automatically detects answers in adjacent cells.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Workbook filename (.xlsx file)"
                }
            },
            "required": ["filename"]
        }
    },
    {
        "name": "detect_hidden_content",
        "description": "List all hidden rows, columns, and worksheets. Shows which content is hidden in the workbook.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Workbook filename (.xlsx file)"
                }
            },
            "required": ["filename"]
        }
    },
    {
        "name": "comprehensive_analysis",
        "description": "Full workbook analysis - extracts comments, questions, hidden content, and provides summary statistics.",
        "inputSchema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "Workbook filename (.xlsx file)"
                }
            },
            "required": ["filename"]
        }
    }
]

def handle_request(request: Dict[str, Any]) -> Dict[str, Any]:
    """Handle MCP JSON-RPC request"""
    method = request.get("method")
    params = request.get("params", {})
    req_id = request.get("id", 0)

    try:
        # Handle MCP protocol methods
        if method == "initialize":
            return {
                "jsonrpc": "2.0",
                "id": req_id,
                "result": {
                    "protocolVersion": "2024-11-05",
                    "capabilities": {
                        "tools": {}
                    },
                    "serverInfo": {
                        "name": "excel-python-mcp",
                        "version": "1.0.0"
                    }
                }
            }

        elif method == "tools/list":
            return {
                "jsonrpc": "2.0",
                "id": req_id,
                "result": {
                    "tools": TOOLS
                }
            }

        elif method == "tools/call":
            tool_name = params.get("name")
            arguments = params.get("arguments", {})
            filename = arguments.get("filename")

            # Get file buffer from request if provided (base64 encoded)
            file_buffer = None
            if "fileContent" in request:
                file_buffer = base64.b64decode(request["fileContent"])

            # Call appropriate analyzer method
            if tool_name == "extract_comments":
                result = ExcelAnalyzer.extract_comments(filename, file_buffer)
            elif tool_name == "extract_questions":
                result = ExcelAnalyzer.extract_questions(filename, file_buffer)
            elif tool_name == "detect_hidden_content":
                result = ExcelAnalyzer.detect_hidden_content(filename, file_buffer)
            elif tool_name == "comprehensive_analysis":
                result = ExcelAnalyzer.comprehensive_analysis(filename, file_buffer)
            else:
                raise ValueError(f"Unknown tool: {tool_name}")

            return {
                "jsonrpc": "2.0",
                "id": req_id,
                "result": {
                    "content": json.dumps(result, indent=2)
                }
            }

        else:
            raise ValueError(f"Unknown method: {method}")

    except Exception as e:
        return {
            "jsonrpc": "2.0",
            "id": req_id,
            "error": {
                "code": -32603,
                "message": "Internal error",
                "data": str(e)
            }
        }

def main():
    """Main server loop - read JSON-RPC from stdin, write to stdout"""
    print("Python Excel Analysis MCP Server starting...", file=sys.stderr)

    for line in sys.stdin:
        if not line.strip():
            continue

        try:
            request = json.loads(line)
            response = handle_request(request)
            print(json.dumps(response), flush=True)
        except Exception as e:
            error_response = {
                "jsonrpc": "2.0",
                "id": 0,
                "error": {
                    "code": -32700,
                    "message": "Parse error",
                    "data": str(e)
                }
            }
            print(json.dumps(error_response), flush=True)

if __name__ == "__main__":
    main()
